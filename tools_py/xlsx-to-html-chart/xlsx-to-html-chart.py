#!/usr/bin/env python3

import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import plotly.graph_objects as go
import openpyxl.chart.bar_chart
import openpyxl.chart.pie_chart
from openpyxl.chart.pie_chart import DoughnutChart
import openpyxl.chart.line_chart

logShow = False  # Set to True to enable debug logging

def log(*args, **kwargs):
    """Simple logging function for debugging."""
    if logShow:
        print(*args, file=sys.stderr, **kwargs)

def extract_chart_title(chart, sheet):
    """
    Extracts the title of a chart, handling both directly typed titles
    and titles linked to a cell.
    """
    title = chart.title
    if title is None:
        return None

    # 1. Check for linked title (strRef)
    # This is often used when the title in Excel is set to reference a cell.
    if hasattr(title, 'strRef') and hasattr(title.strRef, 'f') and title.strRef.f:
        try:
            ref_full = title.strRef.f
            cell_ref_only = ref_full.split('!')[-1].replace('$', '') # Get A1 from Sheet1!$A$1
            
            # For simplicity, we assume the linked cell is on the current sheet.
            # If a title could link to another sheet, you'd need 'wb' here
            # and potentially parse the sheet name from ref_full.
            linked_cell_value = sheet[cell_ref_only].value
            if linked_cell_value is not None:
                log(f"  Debug: Found linked title '{linked_cell_value}' from '{ref_full}'")
                return str(linked_cell_value).strip()
        except Exception as e:
            log(f"  Warning: Could not parse linked chart title from strRef '{title.strRef.f}': {e}")
            # Fall through to rich text parsing if linked title extraction fails

    # 2. Check for direct title (rich text)
    # This is used when the title in Excel is typed directly into the chart.
    try:
        if hasattr(title, 'tx') and hasattr(title.tx, 'rich') and hasattr(title.tx.rich, 'p'):
            # The structure for rich text titles can be nested.
            # We look for the 't' attribute within 'r' elements.
            if title.tx.rich.p and len(title.tx.rich.p) > 0:
                # Concatenate text from all runs in the first paragraph
                full_text = []
                for run in title.tx.rich.p[0].r:
                    if hasattr(run, 't') and run.t is not None:
                        full_text.append(str(run.t))
                if full_text:
                    direct_title = "".join(full_text).strip()
                    log(f"  Debug: Found direct title '{direct_title}'")
                    return direct_title
        return None
    except Exception as e:
        log(f"  Warning: Could not parse direct chart title (rich text): {e}")
        return None

def parse_range(sheet, ref):
    """Parses an Excel range reference and extracts values from the sheet."""
    if '!' in ref:
        ref = ref.split('!')[1]
    ref = ref.replace('$', '')

    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except Exception as e:
        log(f"  Warning: Could not parse range '{ref}': {e}")
        return []

    values = []
    # Determine if range is primarily vertical or horizontal to extract values
    if (max_row - min_row + 1) >= (max_col - min_col + 1):  # Vertical (column)
        for row_idx in range(min_row, max_row + 1):
            cell = sheet.cell(row=row_idx, column=min_col)
            values.append(cell.value)
    else:  # Horizontal (row)
        for col_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row=min_row, column=col_idx)
            values.append(cell.value)
    return values

def extract_axis_titles(sheet, chart):
    """Extracts titles for X and Y axes of a chart."""
    category_axis_title = None
    value_axis_title = None

    try:
        if hasattr(chart, 'x_axis') and chart.x_axis and chart.x_axis.title:
            if isinstance(chart.x_axis.title, str):
                category_axis_title = chart.x_axis.title.strip()
            elif hasattr(chart.x_axis.title, 'tx') and hasattr(chart.x_axis.title.tx, 'rich'):
                p = chart.x_axis.title.tx.rich.p
                if p and hasattr(p[0], 'r') and p[0].r and hasattr(p[0].r[0], 't'):
                    category_axis_title = p[0].r[0].t.strip()
            # Handle axis title linked to cell (similar to chart title)
            elif hasattr(chart.x_axis.title, 'strRef') and chart.x_axis.title.strRef and chart.x_axis.title.strRef.f:
                ref_full = chart.x_axis.title.strRef.f
                cell_ref_only = ref_full.split('!')[-1].replace('$', '')
                linked_value = sheet[cell_ref_only].value
                if linked_value is not None:
                    category_axis_title = str(linked_value).strip()

        if hasattr(chart, 'y_axis') and chart.y_axis and chart.y_axis.title:
            if isinstance(chart.y_axis.title, str):
                value_axis_title = chart.y_axis.title.strip()
            elif hasattr(chart.y_axis.title, 'tx') and hasattr(chart.y_axis.title.tx, 'rich'):
                p = chart.y_axis.title.tx.rich.p
                if p and hasattr(p[0], 'r') and p[0].r and hasattr(p[0].r[0], 't'):
                    value_axis_title = p[0].r[0].t.strip()
            # Handle axis title linked to cell
            elif hasattr(chart.y_axis.title, 'strRef') and chart.y_axis.title.strRef and chart.y_axis.title.strRef.f:
                ref_full = chart.y_axis.title.strRef.f
                cell_ref_only = ref_full.split('!')[-1].replace('$', '')
                linked_value = sheet[cell_ref_only].value
                if linked_value is not None:
                    value_axis_title = str(linked_value).strip()

    except Exception as e:
        log(f"  Warning: Could not extract axis titles: {e}")

    return category_axis_title, value_axis_title

def extract_chart_data(filename, sheet_name, chart_title_to_find):
    """
    Extracts data for a specific chart from an Excel workbook.
    """
    if not os.path.isfile(filename):
        sys.exit(0)

    wb = load_workbook(filename, data_only=True)

    if sheet_name not in wb.sheetnames:
        log(f"Error: Sheet '{sheet_name}' not found in workbook.")
        sys.exit(1)

    sheet = wb[sheet_name]
    matched_chart = None

    # Normalize the input chart title for comparison
    normalized_chart_title_to_find = chart_title_to_find.strip()

    for chart in sheet._charts:
        # ✅ IMPORTANT: Pass the 'sheet' object to extract_chart_title
        title = extract_chart_title(chart, sheet)
        log(f"Processing chart: {title or '<No Title>'}")
        
        # Compare normalized titles
        if title and title.strip() == normalized_chart_title_to_find:
            matched_chart = chart
            break

    if not matched_chart:
        log(f"Error: No chart matched the title '{chart_title_to_find}' in sheet '{sheet_name}'")
        return None

    chart = matched_chart

    # Detect chart type
    if isinstance(chart, DoughnutChart):
        chart_type = "doughnut"
    elif isinstance(chart, openpyxl.chart.pie_chart.PieChart):
        chart_type = "pie"
    elif isinstance(chart, openpyxl.chart.bar_chart.BarChart):
        chart_type = "bar"
    elif isinstance(chart, openpyxl.chart.line_chart.LineChart):
        chart_type = "line"
    else:
        chart_type = "unknown"
        log(f"Warning: Unknown chart type detected: {type(chart)}")

    category_axis_title, value_axis_title = extract_axis_titles(sheet, chart)

    barmode = 'group'
    orientation = 'v'
    if chart_type == "bar":
        # 'barDir' can be 'col' (vertical bars) or 'bar' (horizontal bars)
        if getattr(chart, "barDir", None) == "bar":
            orientation = 'h'
        elif getattr(chart, "barDir", None) == "col":
            orientation = 'v'
        
        grouping = getattr(chart, "grouping", None)
        if grouping == "stacked":
            barmode = "stack"
        elif grouping == "percentStacked":
            barmode = "relative"
        log(f"Detected barDir: {chart.barDir}, grouping: {grouping}, orientation: {orientation}, barmode: {barmode}")

    extracted_series = []
    chart_global_categories = []

    # Attempt to get global categories for the chart (usually from x_axis)
    if hasattr(chart, 'x_axis') and chart.x_axis and \
       hasattr(chart.x_axis, 'categories') and chart.x_axis.categories and \
       hasattr(chart.x_axis.categories, 'strRef') and chart.x_axis.categories.strRef:
        cat_ref = chart.x_axis.categories.strRef.f
        chart_global_categories = parse_range(sheet, cat_ref)
        log(f"  Debug: Extracted global categories from x_axis.categories.strRef: {chart_global_categories}")

    # Fallback for categories if not found via x_axis.categories
    if not chart_global_categories and chart.series:
        first_series = chart.series[0]
        # Some charts might have categories linked to the first series' cat reference
        if hasattr(first_series, 'cat') and first_series.cat and \
           hasattr(first_series.cat, 'numRef') and first_series.cat.numRef:
            cat_ref = first_series.cat.numRef.f
            chart_global_categories = parse_range(sheet, cat_ref)
            log(f"  Debug: Extracted global categories from first series.cat.numRef: {chart_global_categories}")
        # Another fallback: infer categories from a column adjacent to values
        elif first_series.val and first_series.val.numRef:
            val_ref = first_series.val.numRef.f.replace('$', '').split('!')[-1]
            min_col, min_row, max_col, max_row = range_boundaries(val_ref)
            if min_col > 1: # If values start from column 2 or later, category might be in column 1
                chart_global_categories = [
                    sheet.cell(row=r, column=min_col - 1).value
                    for r in range(min_row, max_row + 1)
                ]
                log(f"  Debug: Inferred global categories from column left of values: {chart_global_categories}")

    for i, series in enumerate(chart.series, start=1):
        try:
            if series.val is None or series.val.numRef is None:
                log(f"  Warning: Series {i} has no value reference, skipping.")
                continue
            value_ref = series.val.numRef.f
            values = parse_range(sheet, value_ref)
            if not values:
                log(f"  Warning: Series {i} values could not be parsed, skipping.")
                continue

            # Use global categories if available, otherwise try series-specific categories
            categories = chart_global_categories
            if not categories and hasattr(series, 'cat') and series.cat and hasattr(series.cat, 'numRef'):
                categories = parse_range(sheet, series.cat.numRef.f)
                log(f"  Debug: Using series-specific categories for series {i}: {categories}")


            if not categories:
                log(f"  Warning: No categories found for series {i}, skipping.")
                continue

            # Extract series title
            title = f"Series {i}" # Default title
            try:
                # Prioritize series title linked to cell
                if hasattr(series, 'tx') and series.tx and hasattr(series.tx, 'strRef') and series.tx.strRef.f:
                    ref_only = series.tx.strRef.f.split('!')[-1].replace('$', '')
                    series_title_from_cell = sheet[ref_only].value
                    if series_title_from_cell is not None:
                        title = str(series_title_from_cell).strip()
                        log(f"  Debug: Series {i} title from strRef: '{title}'")
                elif series.title and series.title.strRef and series.title.strRef.f:
                    ref_only = series.title.strRef.f.split('!')[-1].replace('$', '')
                    series_title_from_cell = sheet[ref_only].value
                    if series_title_from_cell is not None:
                        title = str(series_title_from_cell).strip()
                        log(f"  Debug: Series {i} title from series.title.strRef: '{title}'")
                # Fallback to direct text if not linked
                elif hasattr(series, 'tx') and series.tx and hasattr(series.tx, 'v'): # Direct string value
                    title = str(series.tx.v).strip()
                    log(f"  Debug: Series {i} title from series.tx.v: '{title}'")
                elif hasattr(series, 'title') and series.title and hasattr(series.title, 'v'): # Direct string value
                    title = str(series.title.v).strip()
                    log(f"  Debug: Series {i} title from series.title.v: '{title}'")
            except Exception as e:
                log(f"  Warning: Could not extract title for series {i} (might be direct text or complex): {e}")
                # Continue with default title or no title if extraction failed

            extracted_series.append({
                "title": title,
                "categories": categories,
                "values": values
            })
            log(f"  Debug: Added series '{title}' with {len(categories)} categories and {len(values)} values.")
        except Exception as e:
            log(f"  Warning: Failed to extract series {i}: {e}")
            continue

    if not extracted_series:
        log(f"Error: No series data extracted for chart '{chart_title_to_find}'.")
        return None

    return {
        "chart": chart_title_to_find, # Use the original chart title passed
        "category_axis_title": category_axis_title,
        "value_axis_title": value_axis_title,
        "series": extracted_series,
        "barmode": barmode,
        "orientation": orientation,
        "chart_type": chart_type,
        "global_categories": chart_global_categories
    }

def normalize_category(cat):
    """Normalizes category values for consistent comparison and display."""
    if isinstance(cat, str):
        return cat.strip()
    return str(cat) if cat is not None else None

def create_plotly_chart(chart_data):
    """Creates an HTML div with a Plotly chart from the extracted data."""
    fig = go.Figure()
    
    # see https://plotly.com/python/reference/layout/
    fig.update_layout(
        autosize = True,
        margin = dict(l=20, r=20, t=80, b=80), # give a bit more horizontal margin because will overlap the right border otherwise (default is 80) 
        paper_bgcolor = 'rgba(0,0,0,0)', # transparent background
        plot_bgcolor = 'rgba(108,117,125,0)', # bootstrap secondary with opacity for plot area
        modebar_bgcolor = 'rgba(0,0,0,0)', # transparent background for toolbox
        modebar_color = 'rgba(108,117,125, 0.5)', # 0.5 opacity for toolbar buttons 
        modebar_activecolor = 'rgba(108,117,125, 1.00)', # full color for active toolbar button
        font_color = 'rgba(108,117,125, 1.00)' # bootstrap secondary without opacity for global fonts
    )

    grid_color = 'rgba(255,193,7,0.5)'
    grid_dash = 'dot'

    fig.update_xaxes(
        gridcolor = grid_color,
        griddash = grid_dash
    )

    fig.update_yaxes(
        gridcolor = grid_color,
        griddash = grid_dash
    )

    chart_type = chart_data.get("chart_type", "bar")
    orientation = chart_data.get("orientation", "v")
    all_categories = [normalize_category(c) for c in chart_data.get("global_categories", [])]

    if chart_type in ("pie", "doughnut"):
        if chart_data['series']:
            series = chart_data['series'][0] # Pie/Doughnut charts typically have one series
            labels = [normalize_category(c) for c in series['categories']]
            pie_params = dict(labels=labels, values=series['values'], name=series['title'])
            if chart_type == "doughnut":
                pie_params["hole"] = 0.4  # Makes it a doughnut chart
            fig.add_trace(go.Pie(**pie_params))
    else:
        for series in chart_data['series']:
            # Create a dictionary for efficient lookup of values by category
            cat_to_val = {
                normalize_category(cat): val
                for cat, val in zip(series['categories'], series['values'])
            }
            # Align values with the global categories to ensure correct plotting
            aligned_y = [cat_to_val.get(cat, None) for cat in all_categories]

            if chart_type == "bar":
                trace_args = {
                    "name": series['title'],
                    "orientation": orientation
                }
                if orientation == 'h': # Horizontal bars: x=values, y=categories
                    trace_args.update(x=aligned_y, y=all_categories)
                else: # Vertical bars: x=categories, y=values
                    trace_args.update(x=all_categories, y=aligned_y)
                fig.add_trace(go.Bar(**trace_args))
            elif chart_type == "line":
                fig.add_trace(go.Scatter(x=all_categories, y=aligned_y, mode="lines+markers", name=series['title']))

        if chart_type == "bar":
            fig.update_layout(barmode=chart_data.get("barmode", "group"))
            # Ensure category order for bar charts
            if orientation == 'v':
                fig.update_xaxes(categoryorder='array', categoryarray=all_categories)
            else:
                fig.update_yaxes(categoryorder='array', categoryarray=all_categories)

    fig.update_layout(
        title=chart_data['chart'], # Use the extracted chart title
        xaxis_title=chart_data.get('category_axis_title'),
        yaxis_title=chart_data.get('value_axis_title'),
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5)
    )
    return fig.to_html(full_html=False, config={'responsive': True, 'displaylogo': False })

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python extract_chart_data.py <workbook.xlsx> <sheet name> <chart title>", file=sys.stderr)
        sys.exit(1)

    filename = sys.argv[1]
    sheet_name = sys.argv[2]
    chart_title = sys.argv[3]

    chart_data = extract_chart_data(filename, sheet_name, chart_title)
    if chart_data is None:
        sys.exit(1)

    html = create_plotly_chart(chart_data)
    print(html)