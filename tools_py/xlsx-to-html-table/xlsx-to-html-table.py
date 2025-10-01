#!/usr/bin/env python3

import sys
import html
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from zipfile import BadZipFile

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def table_range_to_html(ws, cell_range):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    html_rows = ['<table cellspacing="0" cellpadding="4" style="border-collapse: collapse; border: 1px solid #000;">']

    # Map merged cells
    merged_cells = ws.merged_cells.ranges
    merged_map = {}
    for merged in merged_cells:
        m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged))
        for row in range(m_min_row, m_max_row + 1):
            for col in range(m_min_col, m_max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                merged_map[coord] = {
                    "top_left": (row == m_min_row and col == m_min_col),
                    "rowspan": m_max_row - m_min_row + 1,
                    "colspan": m_max_col - m_min_col + 1
                }

    def is_row_hidden(row_idx):
        dim = ws.row_dimensions.get(row_idx)
        return dim.hidden if dim and dim.hidden else False

    def wrap_rotated_content(cell, content):
        rot = cell.alignment.text_rotation
        if rot in [None, 0]:
            return content

        if rot == 255:
            # Excel's vertical text (top-to-bottom)
            style = (
                "writing-mode: vertical-rl; transform: rotate(180deg); "
                "white-space: nowrap; display: block; height: 100%;"
            )
        else:
            # Fix rotation:
            # Excel negative rotation stored by openpyxl as 90 + abs(neg_value)
            if 90 < rot <= 180:
                css_rot = rot - 90  # e.g. 104 -> 14 deg clockwise
            else:
                css_rot = -rot      # normal positive rotation => invert for CSS

            style = (
                f"transform: rotate({css_rot}deg); "
                "transform-origin: center center; "
                "display: inline-block; white-space: nowrap; "
                "max-width: 100%; overflow: hidden;"
            )

        return f'<div style="{style}">{content}</div>'

    def get_style(cell):
        styles = ['border: 1px solid #000;']

        # Horizontal alignment
        align = cell.alignment.horizontal
        if align:
            styles.append(f'text-align: {align};')

        # Vertical alignment
        valign = cell.alignment.vertical
        if valign:
            styles.append(f'vertical-align: {valign};')

        styles.append('max-width: 150px; max-height: 100px; overflow: hidden;')

        return f' style="{" ".join(styles)}"' if styles else ''

    # Gather visible rows
    visible_rows = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        if not is_row_hidden(row[0].row):
            visible_rows.append(row)

    if not visible_rows:
        return "<!-- No visible rows to render -->"

    # Render <thead>
    html_rows.append('  <thead>')
    html_rows.append('    <tr>')
    for cell in visible_rows[0]:
        coord = cell.coordinate
        if coord in merged_map and not merged_map[coord]["top_left"]:
            continue
        value = cell.value if cell.data_type != "f" else cell._value
        cell_value = html.escape(str(value)) if value is not None else ''
        content = wrap_rotated_content(cell, cell_value)
        attrs = get_style(cell)
        if coord in merged_map:
            info = merged_map[coord]
            if info["rowspan"] > 1:
                attrs += f' rowspan="{info["rowspan"]}"'
            if info["colspan"] > 1:
                attrs += f' colspan="{info["colspan"]}"'
        html_rows.append(f'      <th{attrs}>{content}</th>')
    html_rows.append('    </tr>')
    html_rows.append('  </thead>')

    # Render <tbody>
    html_rows.append('  <tbody>')
    for row in visible_rows[1:]:
        html_rows.append('    <tr>')
        for cell in row:
            coord = cell.coordinate
            if coord in merged_map and not merged_map[coord]["top_left"]:
                continue
            value = cell.value if cell.data_type != "f" else cell._value
            cell_value = html.escape(str(value)) if value is not None else ''
            content = wrap_rotated_content(cell, cell_value)
            attrs = get_style(cell)
            if coord in merged_map:
                info = merged_map[coord]
                if info["rowspan"] > 1:
                    attrs += f' rowspan="{info["rowspan"]}"'
                if info["colspan"] > 1:
                    attrs += f' colspan="{info["colspan"]}"'
            html_rows.append(f'      <td{attrs}>{content}</td>')
        html_rows.append('    </tr>')
    html_rows.append('  </tbody>')
    html_rows.append('</table>')

    return '\n'.join(html_rows)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 xlsx-to-html-table.py <file.xlsx> <range> [sheet_name]", file=sys.stderr)
        print("Example: python3 xlsx-to-html-table.py file.xlsx B2:D10 Sheet1", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    cell_range = sys.argv[2]
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else None

    try:
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        html_output = table_range_to_html(ws, cell_range)
        print(html_output)
    except FileNotFoundError:
        sys.exit(1)
    except BadZipFile:
        print(f"Error: '{file_path}' is not a valid .xlsx file or is corrupted.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
        sys.exit(1)
